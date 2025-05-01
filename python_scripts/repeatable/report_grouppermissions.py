#!/usr/bin/env python
# This script generates a spreadsheet report that lists all the permissions in archivesspace by the first row, with
# each column displaying the permission and each row displaying the user group. If a user group has a permission, it is
# marked with the text of that permission in the spreadsheet or if not, FALSE. This is to check to make sure permissions
# are the same for each user group across all repositories.

import mysql.connector as mysql
import os

from datetime import date
from dotenv import load_dotenv, find_dotenv
from mysql.connector import errorcode
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Font
from pathlib import Path

# Logging
logger.remove()
log_path = Path('../../logs', 'report_grouppermissions_{time:YYYY-MM-DD}.log')
logger.add(str(log_path), format="{time}-{level}: {message}")

# Find  and load environment-specific .env file
env_file = find_dotenv(f'.env.{os.getenv("ENV", "dev")}')
load_dotenv(env_file)

class ASpaceDatabase:

    def __init__(self, as_db_un, as_db_pw, as_db_host, as_db_name, as_db_port):
        """
        Handles the connection to and data retrieval from the ArchivesSpace database

        Args:
            as_db_un (str): the username for the account to connect to the ArchivesSpace database
            as_db_pw (str): the password for the account to connect to the ArchivesSpace database
            as_db_host (str): the hostname for the ArchivesSpace database
            as_db_name (str): the name of the ArchivesSpace database
            as_db_port (int): the port number of the ArchivesSpace database
        """
        self.aspace_username = as_db_un
        self.aspace_password = as_db_pw
        self.aspace_host = as_db_host
        self.aspace_name = as_db_name
        self.aspace_port = as_db_port
        self.connection, self.cursor = self.connect_db()

    def connect_db(self):
        """
        Connects to the ArchivesSpace test database with credentials provided in local secrets.py file

        Returns:
             test_connect (mysql.connection): The connection to the database
             test_cursor (mysql.connection.cursor): The cursor of results for the database
        """

        try:
            self.connection = mysql.connect(user=self.aspace_username,
                                            password=self.aspace_password,
                                            host=self.aspace_host,
                                            database=self.aspace_name,
                                            port=self.aspace_port)
        except mysql.Error as error:
            if error.errno == errorcode.ER_ACCESS_DENIED_ERROR:
                record_error('connect_db() - Failed to authorize username/password', error)
                raise error
            elif error.errno == errorcode.ER_BAD_DB_ERROR:
                record_error('connect_db() - Database does not exist', error)
                raise error
            else:
                record_error('connect_db() - Other error when connecting to the database', error)
                raise error
        else:
            self.cursor = self.connection.cursor()
            return self.connection, self.cursor


    def query_database(self, statement):
        """
        Runs a query on the database

        Args:
            statement (str): The MySQL statement to run against the database

        Returns:
            worksheet (openpysl.worksheet): An openpyxl worksheet class
        """
        try:
            self.cursor.execute(statement)
        except mysql.Error as error:
            record_error('query_database() - SQL query was invalid', error)
            raise error
        else:
            results = self.cursor.fetchall()
        return results

    def close_connection(self):
        """
        Closes the cursor and connection to the ArchivesSpace database
        """
        self.cursor.close()
        self.connection.close()


class Spreadsheet:

    def __init__(self, spreadsheet_filepath):
        """
        Generate and edit a working spreadsheet for reporting

        Args:
            spreadsheet_filepath (str): the filepath of the spreadsheet generated
        """
        self.spreadsheet_filepath = spreadsheet_filepath
        """self.spreadsheet_filepath (str): The filepath of the spreadsheet being generated"""
        self.wb = self.generate_workbook()
        """self.wb (openpyxl.Workbook): The openpyxl workbook instance of the spreadsheet being generated"""
        self.sheets = {}
        """self.sheets (dict): A dictionary of all the sheets within this instance's workbook (self.wb). key = 
        sheet-name, value = openpysl.worksheet"""


    def generate_workbook(self):
        """
        Creates a new spreadsheet for the data audit output, distinguished by date appended to end of filename

        Returns:
            wb (openpyxl.Workbook): The openpyxl workbook of the spreadsheet being generated for the data audit
            data_worksheet (str): The filepath of the data audit worksheet
        """
        wb = Workbook()
        try:
            wb.save(self.spreadsheet_filepath)
        except FileNotFoundError as invalid_file:
            record_error('generate_workbook() - Unable to save workbook with given filepath', invalid_file)
            raise invalid_file
        return wb

    def create_sheet(self, sheetname):
        """
        Takes a name of a sheet to generate within the workbook (self.wb) for this instance

        Args:
            sheetname (str): the name of the individual sheet within the workbook (self.wb)

        Returns:
            worksheet (openpysl.worksheet): An openpyxl worksheet class
        """
        try:
            worksheet = self.wb.create_sheet(sheetname)
            worksheet.title = sheetname
        except ValueError as bad_sheettitle:
            record_error('create_sheet() - The sheet name provided was not acceptable as a title',
                         bad_sheettitle)
            raise bad_sheettitle
        else:
            self.wb.save(self.spreadsheet_filepath)
            self.sheets[sheetname] = worksheet
            return worksheet


    def write_column_data(self, worksheet, permission, cell_column, cell_row, header=False):
        """
        Takes a list of strings and writes them to a specified column of the given worksheet

        Args:
            worksheet (openpysl.worksheet): An openpyxl worksheet class for the worksheet you want to write to
            permission (str): permission value to write to specific cell value
            cell_column (int): the minimum column to write to, max will not be supplied
            cell_row (int): the minimum row to write to, max will not be supplied
            header (bool): if True, format the permission string with bold and underline
        """
        # for row in worksheet.iter_cols(min_col=min_column, min_row=min_row):
        #     for cell in row:
        worksheet.cell(row=cell_row, column=cell_column).value=permission
        if header is True:
            worksheet.cell(row=cell_row, column=cell_column).font=Font(bold=True, underline='single')
        self.wb.save(self.spreadsheet_filepath)


def record_error(message, status_input):
    """
    Prints and logs an error message and the code/parameters causing the error
    Args:
        message (str): message to prefix the error code
        status_input (str, tuple, bool): error code or input parameters producing the error
    """
    try:
        print(f'{message}: {status_input}')
        logger.error(f'{message}: {status_input}')
    except TypeError as input_error:
        print(f'record_error() - Input is invalid for recording error: {input_error}')
        logger.error(f'record_error() - Input is invalid for recording error: {input_error}')


def main():
    aspace_db = ASpaceDatabase(os.getenv('DB_UN'), os.getenv('DB_PW'), os.getenv('DB_HOST'), os.getenv('DB_NAME'),
                               int(os.getenv('DB_PORT')))
    report_spreadsheet = Spreadsheet(str(Path('../../test_data',
                                              f'report_grouppermissions_{str(date.today())}.xlsx')))
    report_spreadsheet.wb.remove(report_spreadsheet.wb['Sheet'])
    grouppermission_sheet = report_spreadsheet.create_sheet('group_permissions')

    user_group_query = ('SELECT '
                            'permission.description, repo.repo_code, grp.group_code '
                        'FROM '
                            'group_permission AS gp '
                        'JOIN '
                            'permission ON permission.id = gp.permission_id '
                        'JOIN '
                            '`group` AS grp ON grp.id = gp.group_id '
                        'JOIN '
                            'repository AS repo ON repo.id = grp.repo_id '
                        'WHERE '
                            'level != "global" '
                        'ORDER BY '
                            'group_id')

    all_permissions_query = ('SELECT '
                                 '`description` '
                             'FROM '
	                             'permission '
                             'WHERE level != "global"')

    user_groups = aspace_db.query_database(user_group_query)
    all_permissions = aspace_db.query_database(all_permissions_query)

    # Format permissions as a list of strings
    cleaned_permissions = []
    for result in all_permissions:
        cleaned_permissions.append(result[0])

    # Write user_group names as <repository_code>--<user_group_name> as keys and all their permissions as a list in
    # group_permissions dict
    groups_permissions = {}
    group_name = ''
    for user_group in user_groups:
        if str(user_group[1]) + '--' + str(user_group[2]) not in groups_permissions:
            group_name = str(user_group[1]) + '--' + str(user_group[2])
            groups_permissions[group_name] = []
            groups_permissions[group_name].append(user_group[0])
        else:
            groups_permissions[group_name].append(user_group[0])

    # Insert FALSE into permissions list if any of the cleaned_permissions do not exist in said list
    permission_index = 0
    for permission in cleaned_permissions:
        for group_permissions in groups_permissions.values():
            if permission not in group_permissions:
                group_permissions.insert(permission_index, 'FALSE')
        permission_index += 1

    # Write permissions and groups to spreadsheet
    allgroups_row = 2
    for group, permissions in groups_permissions.items():
        report_spreadsheet.write_column_data(grouppermission_sheet, group, 1, allgroups_row)
        allgroups_col = 2
        for permission in permissions:
            report_spreadsheet.write_column_data(grouppermission_sheet, permission, allgroups_col, allgroups_row)
            allgroups_col += 1
        allgroups_row += 1

    # Write header permissions
    column_index = 2
    for permission in cleaned_permissions:
        report_spreadsheet.write_column_data(grouppermission_sheet, permission, column_index, 1, header=True)
        column_index += 1
    aspace_db.close_connection()

if __name__ == "__main__":
    main()
