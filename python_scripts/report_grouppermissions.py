#!/usr/bin/env python
# This script generates a spreadsheet report that lists all the permissions in archivesspace by the first column, then
# each subsequent column is ever user group in an ArchivesSpace instance. If a user group has a permission, it is marked
# as TRUE in the spreadsheet or if not, FALSE. This is to check to make sure permissions are the same for each user
# group across all repositories.

import mysql.connector as mysql
import openpyxl.utils.exceptions

from datetime import date
from mysql.connector import errorcode
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pathlib import Path
from secrets import *

logger.remove()
log_path = Path('../logs', 'report_grouppermissions_{time:YYYY-MM-DD}.log')
logger.add(str(log_path), format="{time}-{level}: {message}")

class ASpaceDatabase:

    def __init__(self, as_db_un, as_db_pw, as_db_host, as_db_name, as_db_port):
        """
        Handles the connection to and data retrieval from the ArchivesSpace database

        Args:
            as_db_un (str): the username for the account to connect to the ArchivesSpace database
            as_db_pw (str): the password for the account to connect to the ArchivesSpace database
            as_db_host (str): the hostname for the ArchivesSpace database
            as_db_name (str): the name of the ArchivesSpace database
            as_db_port (int): the port number of the ArchivesSpace database
        """
        self.aspace_username = as_db_un
        self.aspace_password = as_db_pw
        self.aspace_host = as_db_host
        self.aspace_name = as_db_name
        self.aspace_port = as_db_port
        self.connection, self.cursor = self.connect_db()

    def connect_db(self):
        """
        Connects to the ArchivesSpace test database with credentials provided in local secrets.py file

        Returns:
             test_connect (mysql.connection): The connection to the database
             test_cursor (mysql.connection.cursor): The cursor of results for the database
        """

        try:
            self.connection = mysql.connect(user=self.aspace_username,
                                            password=self.aspace_password,
                                            host=self.aspace_host,
                                            database=self.aspace_name,
                                            port=self.aspace_port)
        except mysql.Error as error:
            if error.errno == errorcode.ER_ACCESS_DENIED_ERROR:
                record_error('connect_db() - Failed to authorize username/password', error)
                raise error
            elif error.errno == errorcode.ER_BAD_DB_ERROR:
                record_error('connect_db() - Database does not exist', error)
                raise error
            else:
                record_error('connect_db() - Other error when connecting to the database', error)
                raise error
        else:
            self.cursor = self.connection.cursor()
            return self.connection, self.cursor


    def query_database(self, statement):
        """
        Runs a query on the database

        Args:
            statement (str): The MySQL statement to run against the database

        Returns:
            worksheet (openpysl.worksheet): An openpyxl worksheet class
        """
        try:
            self.cursor.execute(statement)
        except mysql.Error as error:
            record_error('query_database() - SQL query was invalid', error)
            raise error
        else:
            results = self.cursor.fetchall()
        return results

    def close_connection(self):
        """
        Closes the cursor and connection to the ArchivesSpace database
        """
        self.cursor.close()
        self.connection.close()


class Spreadsheet:

    def __init__(self, spreadsheet_filepath):
        """
        Generate and edit a working spreadsheet for reporting

        Args:
            spreadsheet_filepath (str): the filepath of the spreadsheet generated
        """
        self.spreadsheet_filepath = spreadsheet_filepath
        """self.spreadsheet_filepath (str): The filepath of the spreadsheet being generated"""
        self.wb = self.generate_workbook()
        """self.wb (openpyxl.Workbook): The openpyxl workbook instance of the spreadsheet being generated"""
        self.sheets = {}
        """self.sheets (dict): A dictionary of all the sheets within this instance's workbook (self.wb). key = 
        sheet-name, value = openpysl.worksheet"""


    def generate_workbook(self):
        """
        Creates a new spreadsheet for the data audit output, distinguished by date appended to end of filename

        Returns:
            wb (openpyxl.Workbook): The openpyxl workbook of the spreadsheet being generated for the data audit
            data_worksheet (str): The filepath of the data audit worksheet
        """
        wb = Workbook()
        try:
            wb.save(self.spreadsheet_filepath)
        except FileNotFoundError as invalid_file:
            record_error('generate_workbook() - Unable to save workbook with given filepath', invalid_file)
            raise invalid_file
        return wb

    def create_sheet(self, sheetname):
        """
        Takes a name of a sheet to generate within the workbook (self.wb) for this instance

        Args:
            sheetname (str): the name of the individual sheet within the workbook (self.wb)

        Returns:
            worksheet (openpysl.worksheet): An openpyxl worksheet class
        """
        try:
            worksheet = self.wb.create_sheet(sheetname)
            worksheet.title = sheetname
        except ValueError as bad_sheettitle:
            record_error('create_sheet() - The sheet name provided was not acceptable as a title',
                         bad_sheettitle)
            raise bad_sheettitle
        else:
            self.wb.save(self.spreadsheet_filepath)
            self.sheets[sheetname] = worksheet
            return worksheet


    def write_headers(self, worksheet, headers): # TODO: write unittests for this function
        """
        Takes a list of strings and writes them to the top row for a sheet in the data audit spreadsheet

        Args:
            worksheet (openpysl.worksheet): An openpyxl worksheet class for the worksheet you want to write to
            headers (list): List of strings to be headers on the top row of the sheet

        Returns:
            worksheet (openpysl.worksheet): An openpyxl worksheet class
        """
        header_index = 0
        for row in worksheet.iter_rows(min_row=1, max_col=len(headers)):
            for cell in row:
                worksheet[cell.coordinate] = headers[header_index]
                worksheet[cell.coordinate].font = Font(bold=True, underline='single')
                header_index += 1
        self.wb.save(self.spreadsheet_filepath)


def record_error(message, status_input):
    """
    Prints and logs an error message and the code/parameters causing the error
    Args:
        message (str): message to prefix the error code
        status_input (str, tuple, bool): error code or input parameters producing the error
    """
    try:
        print(f'{message}: {status_input}')
        logger.error(f'{message}: {status_input}')
    except TypeError as input_error:
        print(f'record_error() - Input is invalid for recording error: {input_error}')
        logger.error(f'record_error() - Input is invalid for recording error: {input_error}')


def main():
    aspace_db = ASpaceDatabase(as_dbstag_un, as_dbstag_pw, as_dbstag_host, as_dbstag_database, as_dbstag_port)
    report_spreadsheet = Spreadsheet(str(Path('../test_data',
                                              f'report_grouppermissions_{str(date.today())}.xlsx')))
    report_spreadsheet.wb.remove(report_spreadsheet.wb['Sheet'])
    grouppermission_sheet = report_spreadsheet.create_sheet('group_permissions')
    user_group_query = ('SELECT '
                            'permission.description, repo.repo_code, grp.group_code '
                        'FROM '
                            'group_permission AS gp '
                        'JOIN '
                            'permission ON permission.id = gp.permission_id '
                        'JOIN '
                            '`group` AS grp ON grp.id = gp.group_id '
                        'JOIN '
                            'repository AS repo ON repo.id = grp.repo_id '
                        'ORDER BY '
                            'group_id')
    all_permissions_query = ('SELECT '
                                 '`description` '
                             'FROM '
	                             'permission ')
    user_groups = aspace_db.query_database(user_group_query)
    all_permissions = aspace_db.query_database(all_permissions_query)
    cleaned_permissions = []
    for result in all_permissions:
        cleaned_permissions.append(result[0])
    groups_permissions = {}
    group_name = ''
    for user_group in user_groups:
        if str(user_group[1]) + '--' + str(user_group[2]) not in groups_permissions:
            group_name = str(user_group[1]) + '--' + str(user_group[2])
            groups_permissions[group_name] = []
            groups_permissions[group_name].append(user_group[0])
        else:
            groups_permissions[group_name].append(user_group[0])
    print(groups_permissions['_archivesspace--searchindex'])
    aspace_db.close_connection()

if __name__ == "__main__":
    main()
