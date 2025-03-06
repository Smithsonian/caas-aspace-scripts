#!/usr/bin/env python

# This script consists of unittests for report_grouppermissions.py
import contextlib
import io
import unittest

import mysql.connector
import openpyxl
import openpyxl.utils.exceptions

from python_scripts.repeatable.report_grouppermissions import *

env_file = find_dotenv(f'.env.dev')
load_dotenv(env_file)
test_dbconnection = ASpaceDatabase(os.getenv('DB_UN'), os.getenv('DB_PW'), os.getenv('DB_HOST'), os.getenv('DB_NAME'),
                               int(os.getenv('DB_PORT')))

class TestASpaceDatabaseClass(unittest.TestCase):

    def test_good_connection(self):
        """
        Tests connection with good credentials as founds within the secrets.py file
        """
        self.assertIsNotNone(test_dbconnection.connection)
        self.assertIsNotNone(test_dbconnection.cursor)

    def test_bad_connection(self):
        """
        Tests that an error is raised with making a connection with a bad username and password using the database found
        in the secrets.py file
        """
        with self.assertRaises(mysql.Error):
            ASpaceDatabase('bad_un', 'bad_pw', 'bad_host', os.getenv('DB_NAME'),
                               int(os.getenv('DB_PORT')))

    def test_good_query(self):
        """
        Tests that the query_database() function returns a list of users from an ASpace database
        """
        test_query = ('SELECT name, username FROM user')
        test_results = test_dbconnection.query_database(test_query)
        self.assertIsNotNone(test_results)
        self.assertIsInstance(test_results, list)

    def test_bad_query(self):
        """
        Tests that a badly formatted SQL query will raise a mysql.Error
        """
        test_bad_query = ('SELECT nothing, username FROM user')
        with self.assertRaises(mysql.Error):
            test_dbconnection.query_database(test_bad_query)

class TestSpreadsheetClass(unittest.TestCase):

    def test_generate_workbook_good(self):
        """
        Tests generating an openpyxl Workbook instance and assigns it as a Spreadsheet instance variable
        """
        test_spreadsheet_filepath = str(Path('../test_data', f'test_report_grouppermissions_{str(date.today())}.xlsx'))
        test_spreadsheet = Spreadsheet(test_spreadsheet_filepath)
        self.assertIsInstance(test_spreadsheet, Spreadsheet)
        self.assertIsInstance(test_spreadsheet.wb, openpyxl.Workbook)
        os.remove(test_spreadsheet_filepath)

    def test_generate_workbook_bad(self):
        """
        Tests generating an openpyxl Workbook instance with a bad filepath to make sure it raises an error
        """
        test_spreadsheet_filepath = str(f'//bad_filepath')
        with self.assertRaises(FileNotFoundError):
            Spreadsheet(test_spreadsheet_filepath)

    def test_create_sheet(self):
        """
        Tests creating a sheet within an existing spreadsheet
        """
        test_spreadsheet_filepath = str(Path('../test_data', f'test_report_grouppermissions_{str(date.today())}.xlsx'))
        test_spreadsheet = Spreadsheet(test_spreadsheet_filepath)
        test_sheetname = 'test'
        test_spreadsheet.create_sheet(test_sheetname)

        test_sheetnames = test_spreadsheet.wb.sheetnames
        self.assertIn("test", test_sheetnames)
        self.assertIn('test', test_spreadsheet.sheets)
        test_spreadsheet.wb.close()
        os.remove(test_spreadsheet_filepath)

    def test_create_sheet_error(self):
        """
        Tests creating a bad sheet using a bad sheet name to raise a ValueError
        """
        test_spreadsheet_filepath = str(Path('../test_data', f'test_report_grouppermissions_{str(date.today())}.xlsx'))
        test_spreadsheet = Spreadsheet(test_spreadsheet_filepath)
        test_sheetname = 'my:test:sheet'
        with self.assertRaises(ValueError):
            test_spreadsheet.create_sheet(test_sheetname)
        test_spreadsheet.wb.close()
        os.remove(test_spreadsheet_filepath)


    def test_write_column_data(self):
        """
        Test write_column_data() function by writing a list of test data to a spreadsheet and assert that those test
        data values exist in the spreadsheet.
        """
        test_perms = {'_archivesspace--searchindex': ["test_header_1", "test_header_2", "test_header_3"],
                        '_archivesspace--administrator': ["test_header_1", "test_header_2", "test_header_3"]}
        test_spreadsheet_filepath = str(Path('../test_data', f'test_report_grouppermissions_{str(date.today())}.xlsx'))
        test_spreadsheet = Spreadsheet(test_spreadsheet_filepath)
        test_worksheet = test_spreadsheet.create_sheet('test')
        test_spreadsheet.wb.remove(test_spreadsheet.wb['Sheet'])
        column = 1
        for group, levels in test_perms.items():
            row = 1
            test_spreadsheet.write_column_data(test_worksheet, group, column, row, header=True)
            row += 1
            for permission in levels:
                test_spreadsheet.write_column_data(test_worksheet, permission, column, row)
                row += 1
            column += 1

        if "test" in test_spreadsheet.wb.sheetnames:
            test_sheet = test_spreadsheet.wb["test"]
            column_index = 1
            row_index = 1
            for group, permissions in test_perms.items():
                if row_index == 1:
                    cell = test_sheet.cell(row=row_index, column=column_index)
                    self.assertIn(cell.value, group)
                else:
                    for permission in permissions:
                        self.assertIn(cell.value, permission)
                        row_index += 1
                column_index += 1

        test_spreadsheet.wb.close()
        os.remove(test_spreadsheet_filepath)

class TestRecordError(unittest.TestCase):

    def test_str_input(self):
        """
        Tests that a string input for the record_error() status_input variable properly prints
        """
        f = io.StringIO()
        with contextlib.redirect_stdout(f):
            record_error('This is a test error', 'Error 404 - page not found')
        self.assertTrue(r"""This is a test error: Error 404 - page not found""" in f.getvalue())

    def test_tuple_input(self):
        """
        Tests that a tuple input for the record_error() status_input variable properly prints
        """
        f = io.StringIO()
        with contextlib.redirect_stdout(f):
            record_error('This is a test error', ('test_this', 'and_this'))
        self.assertTrue(r"""This is a test error: ('test_this', 'and_this')""" in f.getvalue())

    def test_bool_input(self):
        """
        Tests that a boolean input for the record_error() status_input variable properly prints
        """
        f = io.StringIO()
        with contextlib.redirect_stdout(f):
            record_error('This is a test error', False)
        self.assertTrue(r"""This is a test error: False""" in f.getvalue())

